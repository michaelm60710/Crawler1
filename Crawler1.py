# -*- coding: utf8 -*-
# coding: utf8
import requests
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook

def get_reason (link):
    res = requests.get(link)
    res.encoding = "utf-8"
    soup = BeautifulSoup(res.text, "lxml")
    choice = 0
    #異動條文及理由
    reason_link = ''
    for link in soup.find_all('a'):
        for test in link.findAll(src="/lglaw/images/yellow_btn01.png"):
            reason_link = 'http://lis.ly.gov.tw' + link.get('href')
            choice = 1
    if (reason_link != ''):
	    #取得理由
	    print ('異動條文及理由')
	    all_text = ''
	    if (reason_link==''):
	    	return all_text
	    res = requests.get(reason_link)
	    res.encoding = "utf-8"
	    soup = BeautifulSoup(res.text, "lxml")
	    tag = ".artipud_RS_2" #理由
	    tag1 = ".artiupd_TH_2"#條文
	    reason_list = []
	    reason2_list = []#條文

	    #print ('start')
	    for drink in soup.select('{}'.format(tag)):
	        reason_list.append(drink.get_text())

	    for drink in soup.select('{}'.format(tag1)):
	        reason2_list.append(drink.get_text())

	    length = len(reason_list)
	    for i in range(length):
	    	all_text += "條文 \n"
	    	all_text += (reason2_list[i] + "\n")
	    	all_text += "理由 \n"
	    	all_text += (reason_list[i] + "\n")

	    return all_text # return 理由

	#異動條文
    reason_link = ''
    for link in soup.find_all('a'):
        for test in link.findAll(src="/lglaw/images/yellow_btn03.png"):
            reason_link = 'http://lis.ly.gov.tw' + link.get('href')
    if (reason_link != ''):
	    #取得理由
	    print ('異動條文')
	    all_text = ''
	    if (reason_link==''):
	    	return all_text
	    res = requests.get(reason_link)
	    res.encoding = "utf-8"
	    soup = BeautifulSoup(res.text, "lxml")
	    tag1 = ".artiupd_TH_2"#條文
	    reason2_list = []#條文

	    for drink in soup.select('{}'.format(tag1)):
	        reason2_list.append(drink.get_text())

	    length = len(reason2_list)
	    for i in range(length):
	    	all_text += "條文 \n"
	    	all_text += (reason2_list[i] + "\n")

	    return all_text # return 條文

 	#廢止理由
    reason_link = ''
    for link in soup.find_all('a'):
        for test in link.findAll(src="/lglaw/images/yellow_btn04.png"):
            reason_link = 'http://lis.ly.gov.tw' + link.get('href')
            choice = 1
    if (reason_link != ''):
	    #取得理由
	    print ('廢止理由')
	    all_text = ''
	    if (reason_link==''):
	    	return all_text
	    res = requests.get(reason_link)
	    res.encoding = "utf-8"
	    soup = BeautifulSoup(res.text, "lxml")
	    tag = ".artipud_RS_2" #理由
	    reason_list = []

	    #print ('start')
	    for drink in soup.select('{}'.format(tag)):
	        reason_list.append(drink.get_text())

	    length = len(reason_list)
	    for i in range(length):
	    	all_text += "理由 \n"
	    	all_text += (reason_list[i] + "\n")

	    return all_text # return理由




def get_info(count,output_row,link,ws): #get 理由的網址
    print ('序號: ', count)
    res = requests.get(link)
    res.encoding = "utf-8"
    soup = BeautifulSoup(res.text, "lxml")
    all_text = ''
    time = ''
    count_temp = 0
    #print (soup.text)
    for line in soup.find_all(class_="dettb02"):
        if (len(line.get_text())==1): # wrong catch
            continue
        elif (count_temp==0):
            time = line.get_text()
        else:
            break
        count_temp += 1

    for link in soup.find_all('a'):
        for test in link.findAll(src="/lydb/img/html_icon.png"):
            all_text = get_reason(link.get('href'))

    if (all_text==''): 
    	all_text = 'None'
    '''else :
    	print ('get 理由')'''
    ws.cell(row=output_row+1, column=1).value = count
    ws.cell(row=output_row+1, column=2).value = time
    ws.cell(row=output_row+1, column=3).value = all_text
    
    return ws

def get_firstpage():
	res = requests.get('http://lis.ly.gov.tw/lydbc/lydbkmout')
	res.encoding = "utf-8"
	soup = BeautifulSoup(res.text, "lxml")
	link1 = '' #通過議案 link

	# get the link to 通過議案
	for link in soup.find_all('a'):
	    for test in link.find_all(id="left03"):
	        link1 = link.get('href')
	        
	        
	link1 = 'http://lis.ly.gov.tw' + link1
	print ('first page: ',link1)

	return link1



 
if __name__ == '__main__':
	
	# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
	wb = Workbook()
	ws = wb.get_active_sheet()
	ws.title = 'New Title'  # 设置worksheet的标题
	ws.cell(row=1, column=1).value = "序號"
	ws.cell(row=1, column=2).value = "通過日期"
	ws.cell(row=1, column=3).value = "議案理由"
	

	page = int(sys.argv[2])  
	link1 = sys.argv[1]      #get_firstpage()
	link2 = ''
	title_link = ''
	link_list = []

	
	res = requests.get(link1)
	res.encoding = "utf-8"
	soup = BeautifulSoup(res.text, "lxml")

	#get other pages
	link_list.append(link1)
	for link in soup.find_all('a',class_="linkpage"):
		temp_link = 'http://lis.ly.gov.tw' + link.get('href')
		link_list.append(temp_link)
	link_length = int(len(link_list)/2) + 1
	print ('page: ', page)
	print ('link length: ',link_length)

	count = (page-1)*10 
	output_row = 0
	for i in range(link_length):
		#initial soup
		link1 = link_list[i]
		print ('link: ',link1)
		res = requests.get(link1)
		res.encoding = "utf-8"
		soup = BeautifulSoup(res.text, "lxml")
		
		#start to find 理由
		count_title = 0
		for link in soup.find_all('a',title="詳目"):
		    count_title += 1
		    if(count_title%2==0):
		        count += 1
		        output_row += 1
		        title_link = 'http://lis.ly.gov.tw' + link.get('href')
		        #print (title_link)
		        ws = get_info(count ,output_row, title_link,ws)
	
	print ('count: ',count)		
	print ('end')
	
	wb.save(filename='page_'+str(page)+'.xlsx')